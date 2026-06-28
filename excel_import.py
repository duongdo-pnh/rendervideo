"""Excel batch import -> TTS (multi-threaded) -> LatentSync render queue.

Flow:
  1. process_excel(): read .xlsx, validate each row -> (ready_rows, errors). NO side effects.
  2. submit_jobs(): synthesize audio for every ready row IN PARALLEL (ThreadPoolExecutor —
     TTS is I/O-bound network calls), then enqueue each as a render job. The actual video
     render stays SEQUENTIAL (single GPU) — queue_worker.py already serializes that.

Lives at repo ROOT (not inside the latentsync package) so it can `import database as db`
exactly like web_ui.py / queue_worker.py do.

Render config for Excel jobs is fixed to the UI default: model 256 + GFPGAN (mouth).
Output naming keeps the existing convention "<base>__<INTENT>" so Relive Studio still
matches video <-> product by name.
"""
import os
import re
import unicodedata
import uuid
from pathlib import Path

import pandas as pd

import database as db
import tts_db
from tts_voice import normalize_voice
from latentsync.tts.factory import DEFAULT_PROVIDER

ROOT = Path(__file__).parent
TTS_AUDIO_DIR = ROOT / "uploads" / "tts"   # stable dir (uploads/ persists; pipeline temp gets wiped)


# Template đầy đủ (header dòng 1) — khớp file mẫu khách dùng. Hệ CHỈ xử lý các cột:
#   product_name, subText, video_type, question_type, video_path(optional), tts_provider, tts_voice
#   -> render xong ghi ngược 'video_done' (đường dẫn local). Cột khác giữ nguyên, không xử lý.
COLUMNS = ["product_name", "product_link", "product_info", "subText",
           "video_type", "question_type", "video_path", "video_url",
           "video_done", "tts_provider", "tts_voice"]

# Tên cột chấp nhận (alias) khi đọc — tương thích file cũ.
COL_PRODUCT = ("product_name", "product")
COL_TEXT = ("subText", "subtext", "text")

# Giá trị cho dropdown trong file mẫu (data validation) — khách chọn, khỏi gõ sai.
VIDEO_TYPE_OPTIONS = ["gioi_thieu", "tra_loi"]
QUESTION_TYPE_OPTIONS = ["ASK_PRICE", "ASK_QUALITY", "ASK_USAGE", "ASK_STOCK", "ASK_BUY",
                         "ASK_SHIPPING", "ASK_VOUCHER", "ASK_RETURN", "ASK_PRODUCT"]
TTS_PROVIDER_OPTIONS = ["vbee", "ausynclab", "local"]


def _voice_options():
    """Mã giọng cố định của hệ (4 vbee + 4 ausynclab) cho dropdown tts_voice."""
    try:
        from latentsync.tts.vbee import VbeeTTS
        from latentsync.tts.ausynclab import AusynclabTTS
        return [code for _, code in VbeeTTS.CURATED_VOICES] + \
               [code for _, code in AusynclabTTS.CURATED_VOICES]
    except Exception:
        return []

# Intent codes the downstream/Relive system matches on — MUST stay in sync with
# INTENT_CHOICES in web_ui.py ("Mã PHẢI trùng intents bên hệ live (ASK_*)").
# video_type maps to one of these; "gioi_thieu"/intro -> None (video giới thiệu, không __INTENT).
INTENT_MAP = {
    "gioi_thieu": None, "gioithieu": None, "intro": None, "gt": None,
    "gia": "ASK_PRICE", "price": "ASK_PRICE", "hoi_gia": "ASK_PRICE",
    "chat_luong": "ASK_QUALITY", "chatluong": "ASK_QUALITY", "quality": "ASK_QUALITY", "review": "ASK_QUALITY",
    "cach_dung": "ASK_USAGE", "usage": "ASK_USAGE", "huong_dan": "ASK_USAGE",
    "con_hang": "ASK_STOCK", "stock": "ASK_STOCK", "size": "ASK_STOCK", "mau": "ASK_STOCK",
    "mua": "ASK_BUY", "buy": "ASK_BUY", "chot_don": "ASK_BUY", "order": "ASK_BUY",
    "ship": "ASK_SHIPPING", "shipping": "ASK_SHIPPING", "giao_hang": "ASK_SHIPPING", "freeship": "ASK_SHIPPING",
    "voucher": "ASK_VOUCHER", "giam_gia": "ASK_VOUCHER", "khuyen_mai": "ASK_VOUCHER", "discount": "ASK_VOUCHER",
    "doi_tra": "ASK_RETURN", "return": "ASK_RETURN", "bao_hanh": "ASK_RETURN", "warranty": "ASK_RETURN",
    "san_pham": "ASK_PRODUCT", "product": "ASK_PRODUCT", "xem_sp": "ASK_PRODUCT",
}


def _ascii(s):
    """Bỏ dấu + lower, để nhận cả 'Giới thiệu' / 'gioi thieu' / 'Trả lời câu hỏi'.

    Thay đ/Đ -> d/D trước (NFKD KHÔNG tách 'đ' nên nếu không xử lý sẽ mất, vd 'đổi trả' -> 'oi tra')."""
    import unicodedata as _u
    s = (s or "").replace("đ", "d").replace("Đ", "D")
    return _u.normalize("NFKD", s).encode("ascii", "ignore").decode().strip().lower()


def is_intro(video_type, question_type=""):
    """video_type chỉ 2 loại: giới thiệu vs trả lời câu hỏi. True = video giới thiệu."""
    a = _ascii(video_type)
    if "tra" in a and "loi" in a:        # trả lời (câu hỏi)
        return False
    if "gioi" in a and "thieu" in a:     # giới thiệu
        return True
    if a in ("answer", "tl", "tlch", "cau_hoi", "cauhoi"):
        return False
    if a in ("intro", "gt", "introduce", ""):
        return True
    # Không rõ: có loại câu hỏi -> coi là trả lời; ngược lại -> giới thiệu.
    return not bool((question_type or "").strip())


def resolve_intent(question_type):
    """Loại câu hỏi -> mã ASK_* hệ thống (khớp INTENT_CHOICES web_ui). None nếu không xác định.

    Chấp nhận: gõ thẳng 'ASK_*', token tiếng Việt/Anh, hoặc nhãn 'Hỏi giá'... (đã bỏ dấu).
    """
    key = _ascii(question_type).replace(" ", "_").replace("-", "_")
    if not key:
        return None
    if key.upper().startswith("ASK_"):
        return key.upper()
    return INTENT_MAP.get(key)


# Chú thích từng cột (cell comment trong file mẫu). Chỉ các cột hệ DÙNG mới có hướng dẫn rõ.
_COL_HELP = {
    "product_name": "TÊN SẢN PHẨM — đặt tên video. Để TRỐNG = câu trả lời CHUNG (tên '__ASK_*').",
    "product_link": "(Không bắt buộc, hệ không xử lý) link sản phẩm.",
    "product_info": "(Không bắt buộc, hệ không xử lý) mô tả sản phẩm.",
    "subText": "Script cần chuyển thành giọng nói (TTS). BẮT BUỘC.",
    "video_type": "Chọn từ dropdown: gioi_thieu (giới thiệu) hoặc tra_loi (trả lời câu hỏi).",
    "question_type": "CHỈ khi video_type = tra_loi — chọn loại câu hỏi (ASK_*) từ dropdown.",
    "video_path": "Đường dẫn local video avatar (không bắt buộc). TRỐNG = dùng video mặc định.",
    "video_url": "(Hệ không xử lý) — để trống.",
    "video_done": "HỆ TỰ ĐIỀN sau khi render xong (đường dẫn local video). Khách để TRỐNG.",
    "tts_provider": "Chọn từ dropdown: vbee / ausynclab / local. TRỐNG = mặc định.",
    "tts_voice": "Chọn giọng từ dropdown. TRỐNG = giọng mặc định của provider.",
}

# Default render config for jobs created from Excel (matches the web UI defaults).
RENDER_DEFAULTS = dict(model_res="256", guidance=1.5, steps=20, seed=1247,
                       enhance_mouth=1, enhance_region="mouth", out_res="720")

IMPORT_LOG = ROOT / "logs" / "import_tts.log"


def _log_import_error(row, msg):
    """Ghi lỗi TTS lúc import ra file để kiểm tra sau (UI chỉ hiện tạm thời)."""
    try:
        from datetime import datetime
        IMPORT_LOG.parent.mkdir(parents=True, exist_ok=True)
        with open(IMPORT_LOG, "a") as f:
            f.write(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] dòng {row}: {msg}\n")
    except Exception:
        pass


# ----------------------------------------------------------------- template

def _col_letter(name):
    from openpyxl.utils import get_column_letter
    return get_column_letter(COLUMNS.index(name) + 1)


def make_template(path, max_rows=1000):
    """Write an .xlsx template: header + ví dụ + comment + DROPDOWN (data validation).

    Dropdown lấy từ sheet ẩn '_lists' (tránh giới hạn 255 ký tự của list inline) cho
    video_type / question_type / tts_provider / tts_voice -> khách chọn, khỏi gõ sai.
    """
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "import"
    ws.append(COLUMNS)

    def row(**kw):
        return [kw.get(c, "") for c in COLUMNS]
    examples = [
        row(product_name="23525384022", subText="Chào cả nhà, hôm nay shop có sản phẩm cực hot!",
            video_type="gioi_thieu", tts_provider="vbee"),
        row(product_name="23525384022", subText="Dạ giá chỉ 299k ạ, đang sale cực mạnh!",
            video_type="tra_loi", question_type="ASK_PRICE", tts_provider="vbee"),
        row(product_name="set kep toc", subText="Bấm giỏ hàng chốt đơn ngay đi ạ!",
            video_type="tra_loi", question_type="ASK_BUY", tts_provider="local"),
    ]
    for r in examples:
        ws.append(r)

    # comment + độ rộng cột
    for idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.comment = Comment(_COL_HELP.get(col, ""), "LatentSync")
        ws.column_dimensions[get_column_letter(idx)].width = max(16, len(col) + 4)

    # sheet ẩn chứa danh sách hợp lệ cho dropdown
    lists = wb.create_sheet("_lists")
    lists.sheet_state = "hidden"
    options = {"video_type": VIDEO_TYPE_OPTIONS, "question_type": QUESTION_TYPE_OPTIONS,
               "tts_provider": TTS_PROVIDER_OPTIONS, "tts_voice": _voice_options()}
    for col_i, (field, vals) in enumerate(options.items(), start=1):
        letter = get_column_letter(col_i)
        lists.cell(row=1, column=col_i, value=field)
        for r_i, v in enumerate(vals, start=2):
            lists.cell(row=r_i, column=col_i, value=v)
        if not vals:
            continue
        rng = f"_lists!${letter}$2:${letter}${len(vals) + 1}"
        dv = DataValidation(type="list", formula1=rng, allow_blank=True, showDropDown=False)
        col_letter = _col_letter(field)
        dv.add(f"{col_letter}2:{col_letter}{max_rows}")
        ws.add_data_validation(dv)

    wb.save(path)
    return path


# ----------------------------------------------------------------- naming

def _safe_stem(name):
    name = unicodedata.normalize("NFC", (name or "").strip())
    return re.sub(r"[^\w.-]+", "_", name) or "job"


def build_name_excel(product, video_type, question_type):
    """Khớp 100% web_ui.build_name(): '<sản phẩm>__<ASK_*>' / '<sản phẩm>' / (sản phẩm rỗng) '__<ASK_*>'.

    product rỗng + trả lời -> '__ASK_*' = CÂU TRẢ LỜI CHUNG (không gắn sản phẩm cụ thể).
    product rỗng + giới thiệu -> 'video'.
    """
    base = str(product).strip() if product else ""
    if is_intro(video_type, question_type):
        return base or "video"
    intent = resolve_intent(question_type)
    if not intent:
        return base or "video"
    return f"{base}__{intent}" if base else f"__{intent}"


# ----------------------------------------------------------------- read + validate

def _clean(val):
    """NaN / whitespace-only -> None; else stripped str."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    return s or None


def _get(raw, names):
    """Lấy giá trị cột đầu tiên có trong raw theo danh sách alias (vd subText/text)."""
    for n in names:
        if n in raw:
            v = _clean(raw.get(n))
            if v is not None:
                return v
    return None


def process_excel(excel_path, default_video=None, default_provider=None, shopee_item_id=None):
    """Read + validate. Returns (ready_rows, errors). No TTS, no DB writes."""
    df = pd.read_excel(excel_path)
    rows, errors = [], []

    for i, raw in df.iterrows():
        excel_row = i + 2  # 1-based + header
        try:
            video = _clean(raw.get("video_path")) or default_video
            if not video or not os.path.exists(video):
                raise ValueError(f"Video không tồn tại: {video!r}")

            text = _get(raw, COL_TEXT)                                       # subText (alias text)
            if not text:
                raise ValueError("subText rỗng")

            provider = _clean(raw.get("tts_provider")) or default_provider  # None -> factory default
            voice = _clean(raw.get("tts_voice"))                            # None -> provider default
            video_type = _clean(raw.get("video_type")) or "gioi_thieu"
            question_type = _clean(raw.get("question_type")) or ""

            # Tên sản phẩm/video: ưu tiên product_name (alias product) của dòng, else Item ID chung.
            # ĐƯỢC PHÉP rỗng -> câu trả lời CHUNG (tên '__ASK_*', không gắn sản phẩm cụ thể).
            product = _get(raw, COL_PRODUCT) or (str(shopee_item_id).strip() if shopee_item_id else None)

            # Trả lời câu hỏi -> bắt buộc có loại câu hỏi hợp lệ (để dựng <sp>__ASK_*).
            if not is_intro(video_type, question_type):
                if not question_type:
                    raise ValueError("video_type='tra_loi' nhưng thiếu question_type (loại câu hỏi).")
                if resolve_intent(question_type) is None:
                    raise ValueError(f"question_type không hợp lệ: {question_type!r} "
                                     f"(dùng gia/mua/ship/... hoặc mã ASK_*).")

            rows.append({
                "row": excel_row,
                "product": product,            # tên base đã resolve (cột product hoặc item_id chung)
                "video_path": video,
                "video_type": video_type,
                "question_type": question_type,
                "text": text,
                "tts_provider": provider,
                "tts_voice": voice,
                "status": "ready",
            })
        except Exception as e:
            errors.append({"row": excel_row, "error": str(e)})

    return rows, errors


# ----------------------------------------------------------------- enqueue vào TTS queue

def submit_jobs(rows, shopee_item_id=None, progress=None, batch_id=None, excel_path=None):
    """ENQUEUE mỗi dòng vào TTS queue bền vững (tts_jobs). KHÔNG gọi TTS tại chỗ —
    tts_worker.py sẽ synth (rate-limit + retry + adaptive throttle) rồi tạo render job.

    excel_path: file Excel gốc của batch -> lưu để sau xuất lại (điền video_done).
    Trả (enqueued, warnings, batch_id, skipped). Voice chuẩn hoá trước (voice guard).
    """
    db.init_db()
    tts_db.init_db()
    TTS_AUDIO_DIR.mkdir(parents=True, exist_ok=True)
    if batch_id is None:
        batch_id = uuid.uuid4().hex[:12]
    tts_db.save_batch(batch_id, excel_path)

    enqueued, warnings, skipped = [], [], []
    dedup = os.getenv("TTS_DEDUP", "1") != "0"      # chống trùng (tắt bằng TTS_DEDUP=0)
    for n, row in enumerate(rows, 1):
        provider = row["tts_provider"] or DEFAULT_PROVIDER
        voice, warn = normalize_voice(provider, row["tts_voice"])
        if warn:
            warnings.append({"row": row["row"], "warn": warn})
            _log_import_error(row["row"], f"voice guard: {warn}")
        # Chống trùng: cùng (sản phẩm+loại+text) đã có job đang chờ/đang chạy/đã xong -> bỏ qua.
        if dedup:
            dup = tts_db.find_duplicate(row["product"], row["video_type"],
                                        row["question_type"], row["text"])
            if dup:
                skipped.append({"row": row["row"], "dup_id": dup,
                                "name": build_name_excel(row["product"], row["video_type"],
                                                         row["question_type"])})
                continue
        tid = tts_db.enqueue(batch_id, row["row"], row["text"], provider, voice,
                             row["product"], row["video_path"], row["video_type"],
                             row["question_type"])
        enqueued.append({"tts_id": tid, "row": row["row"], "text": row["text"][:40]})
        if progress:
            progress(n, len(rows), f"Enqueue {n}/{len(rows)}")

    return enqueued, warnings, batch_id, skipped


# ----------------------------------------------------------------- ghi ngược video_done -> Excel

def export_results(batch_id, out_path=None):
    """Mở Excel gốc của batch, điền cột 'video_done' = đường dẫn local video đã render xong,
    lưu thành file mới. Dòng chưa render xong để trống (hoặc ghi trạng thái).

    Trả (out_path, filled, total). filled = số dòng đã điền được link video.
    """
    from openpyxl import load_workbook

    src = tts_db.get_batch_excel(batch_id)
    if not src or not os.path.exists(src):
        raise FileNotFoundError(f"Không tìm thấy Excel gốc của batch {batch_id}: {src!r}")

    wb = load_workbook(src)
    ws = wb["import"] if "import" in wb.sheetnames else wb.active

    # map tên cột (header dòng 1) -> chỉ số cột
    header = {}
    for c in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=c).value
        if name:
            header[str(name).strip()] = c
    # đảm bảo có cột video_done (thêm vào cuối nếu thiếu)
    done_col = header.get("video_done")
    if not done_col:
        done_col = ws.max_column + 1
        ws.cell(row=1, column=done_col, value="video_done")

    results = tts_db.results_for_batch(batch_id)
    # fallback theo TÊN cho dòng bị dedup bỏ qua (không có tts_job trong batch)
    name_out = {}
    for j in db.list_done_jobs(limit=2000):
        if j.get("output_path"):
            name_out[j["name"]] = j["output_path"]      # list_done_jobs mới nhất trước -> không ghi đè

    def cell(r, field):
        ci = header.get(field)
        return _clean(ws.cell(row=r, column=ci).value) if ci else None

    filled = total = 0
    for r in range(2, ws.max_row + 1):
        # bỏ dòng trống hoàn toàn
        if not any(_clean(ws.cell(row=r, column=c).value) for c in header.values()):
            continue
        total += 1
        out = None
        res = results.get(r)
        if res and res.get("render_status") == "done" and res.get("output_path"):
            out = res["output_path"]
        else:
            # fallback theo tên (dòng dedup-skip hoặc chưa join được)
            product = cell(r, "product_name") or cell(r, "product")
            nm = build_name_excel(product, cell(r, "video_type") or "gioi_thieu",
                                  cell(r, "question_type") or "")
            out = name_out.get(nm)
        if out:
            ws.cell(row=r, column=done_col, value=out)
            filled += 1

    if out_path is None:
        p = Path(src)
        out_path = str(p.with_name(f"{p.stem}_ketqua{p.suffix}"))
    wb.save(out_path)
    return out_path, filled, total
