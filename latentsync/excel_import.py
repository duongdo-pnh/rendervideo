"""ExcelImporter — đọc/ghi sheet "FILE IMPORT" cho batch render pipeline.

Định dạng cố định:
  Row 3 = header, data từ row 5.
  Input cols:  product_name, shopee_item_id, video_path, video_type, text,
               question_type, tts_provider, tts_voice, note
  Output cols: video_done, video_url, status, render_time_s
"""
import os
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "FILE IMPORT"
HEADER_ROW = 3
DATA_START_ROW = 5

VALID_VIDEO_TYPES = {"gioi_thieu", "gia", "mua", "size", "ship", "chat_luong", "con_hang"}


class ExcelImporter:
    def __init__(self, excel_path):
        self.path = str(excel_path)
        self._load()

    # ---------------------------------------------------------------- internal

    def _load(self):
        self.wb = load_workbook(self.path)
        # Case-insensitive sheet lookup
        match = next(
            (s for s in self.wb.sheetnames if s.strip().upper() == "FILE IMPORT"),
            None,
        )
        if match is None:
            raise ValueError(
                f"Không tìm thấy sheet 'FILE IMPORT'. Có: {self.wb.sheetnames}"
            )
        self.ws = self.wb[match]
        self._col_map = self._detect_columns()

    def _detect_columns(self):
        """Scan header row → {normalised_name: col_idx (1-based)}."""
        col_map = {}
        for cell in self.ws[HEADER_ROW]:
            if cell.value:
                name = str(cell.value).strip().lower().replace(" ", "_")
                col_map[name] = cell.column
        return col_map

    def _get_col(self, name):
        return self._col_map.get(name.lower().strip())

    def _cell_val(self, row, col_name):
        col = self._get_col(col_name)
        if col is None:
            return None
        val = self.ws.cell(row=row, column=col).value
        if val is None:
            return None
        s = str(val).strip()
        return s or None

    def _row_is_empty(self, r):
        return all(
            self.ws.cell(row=r, column=c).value is None
            for c in range(1, self.ws.max_column + 1)
        )

    # ---------------------------------------------------------------- public API

    def parse(self):
        """Return list[dict].  Each dict keys: _row, _error, + all column names.
        Rows where BOTH text and video_path are blank are silently skipped."""
        rows = []
        for r in range(DATA_START_ROW, self.ws.max_row + 1):
            if self._row_is_empty(r):
                continue

            text = self._cell_val(r, "text")
            video_path = self._cell_val(r, "video_path")

            if not text and not video_path:
                continue

            row = {
                "_row": r,
                "_error": None,
                "product_name": self._cell_val(r, "product_name"),
                "shopee_item_id": self._cell_val(r, "shopee_item_id"),
                "video_path": video_path,
                "video_type": self._cell_val(r, "video_type") or "gioi_thieu",
                "text": text,
                "question_type": self._cell_val(r, "question_type"),
                "tts_provider": self._cell_val(r, "tts_provider"),
                "tts_voice": self._cell_val(r, "tts_voice"),
                "note": self._cell_val(r, "note"),
                "status": self._cell_val(r, "status"),
                "video_done": self._cell_val(r, "video_done"),
                "video_url": self._cell_val(r, "video_url"),
                "render_time_s": self._cell_val(r, "render_time_s"),
            }

            # Validate
            if not text:
                row["_error"] = "Cột 'text' trống"
            elif not video_path:
                row["_error"] = "Cột 'video_path' trống"
            elif not os.path.exists(video_path):
                row["_error"] = f"video_path không tồn tại: {video_path!r}"
            else:
                vtype = row["video_type"].lower()
                if vtype not in VALID_VIDEO_TYPES:
                    row["_error"] = (
                        f"video_type không hợp lệ: {vtype!r}  "
                        f"(hợp lệ: {', '.join(sorted(VALID_VIDEO_TYPES))})"
                    )

            rows.append(row)
        return rows

    def update_row(self, row_idx, **kwargs):
        """Write output columns to row_idx and save immediately.
        Keys not found in the header are silently ignored."""
        for key, val in kwargs.items():
            col = self._get_col(key)
            if col is None:
                continue
            self.ws.cell(row=row_idx, column=col).value = val
        self.wb.save(self.path)

    def get_pending_rows(self):
        """Only rows with status blank or 'pending' (and no parse error)."""
        return [
            r for r in self.parse()
            if not r.get("_error")
            and (r.get("status") or "").strip().lower() in ("", "pending")
        ]

    def reload(self):
        """Reload workbook from disk (picks up changes made by other processes)."""
        self._load()
